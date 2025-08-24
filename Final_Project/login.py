import tkinter
import tkinter.messagebox
import openpyxl
import os
from PIL import Image, ImageTk
import main_page

script_dir = os.path.dirname(os.path.abspath(__file__))
filepath = os.path.join(script_dir, 'data.xlsx')
BACKGROUND_IMAGE_PATH = os.path.join(script_dir, 'download.png')

global_background_image = None
global_background_photo = None
current_window_width = 0
current_window_height = 0

def load_and_resize_background(window, canvas):
    global global_background_image, global_background_photo, current_window_width, current_window_height

    canvas_width = canvas.winfo_width()
    canvas_height = canvas.winfo_height()

    if canvas_width == 0 or canvas_height == 0:
        window.after(100, lambda: load_and_resize_background(window, canvas))
        return

    if global_background_image is None:
        try:
            global_background_image = Image.open(BACKGROUND_IMAGE_PATH)
        except FileNotFoundError:
            tkinter.messagebox.showerror("Error", f"Background image not found: {BACKGROUND_IMAGE_PATH}")
            return
        except Exception as e:
            tkinter.messagebox.showerror("Error", f"Failed to load background image: {e}")
            return

    if canvas_width != current_window_width or canvas_height != current_window_height:
        resized_image = global_background_image.resize((canvas_width, canvas_height), Image.LANCZOS)
        global_background_photo = ImageTk.PhotoImage(resized_image)
        current_window_width = canvas_width
        current_window_height = canvas_height

    canvas.delete("background")
    canvas.create_image(0, 0, image=global_background_photo, anchor="nw", tags="background")
    canvas.lower("background")


if not os.path.exists(filepath):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Nama", "Gmail", "Pendidikan Terakhir", "Nomor Telepon", "Password", "Bidang yang dikuasai", "Project"])
    wb.save(filepath)
else:
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    header = [cell.value for cell in sheet[1]]
    expected_header = ["Nama", "Gmail", "Pendidikan Terakhir", "Nomor Telepon", "Password", "Bidang yang dikuasai", "Project"]
    if header != expected_header:
        try:
            wb.remove(sheet)
        except:
            pass
        sheet = wb.create_sheet("Sheet1")
        sheet.append(expected_header)
        wb.save(filepath)
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=len(expected_header)).value is None:
            sheet.cell(row=row_idx, column=len(expected_header)).value = ""
    wb.save(filepath)

def open_register():
    dashboard.withdraw()
    register_window()

def open_login():
    dashboard.withdraw()
    login_window()

def back_to_dashboard(current_window):
    current_window.destroy()
    dashboard.deiconify()
    dashboard.state('zoomed')
    dashboard.update_idletasks()
    load_and_resize_background(dashboard, dashboard_canvas)


def register_window():
    reg = tkinter.Toplevel()
    reg.title("Register")
    reg.state('zoomed')
    reg.resizable(True, True)
    reg.config(bg="#0077B5")

    canvas = tkinter.Canvas(reg, highlightthickness=0)
    canvas.pack(fill="both", expand=True)

    def on_reg_configure(event=None):
        load_and_resize_background(reg, canvas)
        frame.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

    canvas.bind("<Configure>", on_reg_configure)
    reg.bind("<Visibility>", on_reg_configure)

    frame = tkinter.Frame(canvas, bg="white", bd=2, relief=tkinter.RIDGE)

    tkinter.Label(frame, text="Register", font=("Poppins", 24, "bold"), bg="white").pack(pady=20)

    fields = ["Nama", "Gmail", "Pendidikan Terakhir", "Nomor Telepon", "Password", "Bidang yang dikuasai", "Project"]
    entries = {}

    for field in fields:
        tkinter.Label(frame, text=field, bg="white", font=("Poppins", 12)).pack(pady=(10, 0))
        if field == "Project":
            ent = tkinter.Text(frame, height=5, width=40, font=("Poppins", 12))
            ent.pack(padx=20, pady=5, fill="x")
        else:
            ent = tkinter.Entry(frame, show="*" if field == "Password" else None, font=("Poppins", 12))
            ent.pack(padx=20, pady=5, fill="x")
        entries[field] = ent

    def save_register():
        mandatory_fields = ["Nama", "Gmail", "Pendidikan Terakhir", "Nomor Telepon", "Password", "Bidang yang dikuasai"]

        values = []
        for field in fields:
            if field == "Project":
                project_input = entries[field].get("1.0", tkinter.END).strip()
                projects_list = [p.strip() for p in project_input.split('\n') if p.strip()]
                values.append('\n'.join(projects_list))
            else:
                entry_value = entries[field].get().strip()
                values.append(entry_value)
                if field in mandatory_fields and entry_value == "":
                    tkinter.messagebox.showerror("Error", f"Kolom '{field}' harus diisi!")
                    return
        
        # Check for mandatory fields again after stripping
        for i, field_name in enumerate(fields):
            if field_name in mandatory_fields and values[i] == "":
                tkinter.messagebox.showerror("Error", f"Kolom '{field_name}' harus diisi!")
                return

        # Gmail duplicate check
        new_gmail = values[1].strip().lower() # Assuming Gmail is at index 1
        
        # Load workbook inside the function to ensure it's up-to-date
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        # Get the 'Gmail' column index dynamically
        header_row = [cell.value for cell in sheet[1]]
        try:
            gmail_col_idx = header_row.index("Gmail")
        except ValueError:
            tkinter.messagebox.showerror("Error", "Kolom 'Gmail' tidak ditemukan di file data.xlsx.")
            return

        existing_gmails = [str(row[gmail_col_idx]).strip().lower() for row in sheet.iter_rows(min_row=2, values_only=True)]
        
        if new_gmail in existing_gmails:
            tkinter.messagebox.showerror("Error", "Gmail sudah terdaftar! Gunakan gmail lain.")
            return

        sheet.append(values)
        wb.save(filepath)

        tkinter.messagebox.showinfo("Sukses", "Registrasi Berhasil!")
        reg.destroy()
        dashboard.deiconify()
        dashboard.state('zoomed')
        dashboard.update_idletasks()
        load_and_resize_background(dashboard, dashboard_canvas)


    button_font = ("Poppins", 14)
    tkinter.Button(frame, text="Submit", command=save_register, font=button_font, bg="#66a6ff", fg="white", bd=0, padx=20, pady=10).pack(pady=15)
    tkinter.Button(frame, text="Kembali", command=lambda: back_to_dashboard(reg), font=button_font, bg="gray", fg="white", bd=0, padx=20, pady=10).pack()

def login_window():
    log = tkinter.Toplevel()
    log.title("Login")
    log.state('zoomed')
    log.resizable(True, True)
    log.config(bg="#0077B5")

    canvas = tkinter.Canvas(log, highlightthickness=0)
    canvas.pack(fill="both", expand=True)

    def on_log_configure(event=None):
        load_and_resize_background(log, canvas)
        frame.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

    canvas.bind("<Configure>", on_log_configure)
    log.bind("<Visibility>", on_log_configure)

    frame = tkinter.Frame(canvas, bg="white", bd=2, relief=tkinter.RIDGE)

    tkinter.Label(frame, text="Login", font=("Poppins", 24, "bold"), bg="white").pack(pady=20)

    tkinter.Label(frame, text="Gmail:", bg="white", font=("Poppins", 12)).pack(pady=(10, 0))
    gmail_entry = tkinter.Entry(frame, font=("Poppins", 12))
    gmail_entry.pack(padx=20, pady=5, fill="x")

    tkinter.Label(frame, text="Kata Sandi:", bg="white", font=("Poppins", 12)).pack(pady=(5, 0))
    password_entry = tkinter.Entry(frame, show="*", font=("Poppins", 12))
    password_entry.pack(padx=20, pady=5, fill="x")

    def check_login():
        gmail = gmail_entry.get().strip().lower()
        password = password_entry.get().strip()

        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        found = False
        current_user_gmail = "" # Change to store gmail
        
        header_row = [cell.value for cell in sheet[1]]
        try:
            gmail_col_idx = header_row.index("Gmail")
            password_col_idx = header_row.index("Password")
        except ValueError:
            tkinter.messagebox.showerror("Error", "Kolom 'Gmail' atau 'Password' tidak ditemukan di file data.xlsx.")
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if gmail == str(row[gmail_col_idx]).strip().lower() and password == str(row[password_col_idx]).strip():
                found = True
                current_user_gmail = str(row[gmail_col_idx]) # Store the logged-in user's gmail
                break

        if found:
            tkinter.messagebox.showinfo("Sukses", "Login Berhasil!")
            log.destroy()
            root = tkinter.Toplevel(dashboard)
            root.state('zoomed')
            root.resizable(True, True)
            main_page.MainPage(root, filepath, current_user=current_user_gmail) # Pass gmail
        else:
            tkinter.messagebox.showerror("Error", "Gmail atau Kata Sandi Salah")

    button_font = ("Poppins", 14)
    tkinter.Button(frame, text="Login", command=check_login, font=button_font, bg="#266eeb", fg="white", bd=0, padx=20, pady=10).pack(pady=15)
    tkinter.Button(frame, text="Kembali", command=lambda: back_to_dashboard(log), font=button_font, bg="gray", fg="white", bd=0, padx=20, pady=10).pack()

# Dashboard Window
dashboard = tkinter.Tk()
dashboard.title("Dashboard")
dashboard.state('zoomed')
dashboard.resizable(True, True)
dashboard.config(bg="#0077B5")

dashboard_canvas = tkinter.Canvas(dashboard, highlightthickness=0)
dashboard_canvas.pack(fill="both", expand=True)

def configure_dashboard_canvas(event):
    load_and_resize_background(dashboard, dashboard_canvas)
    frame.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

dashboard_canvas.bind("<Configure>", configure_dashboard_canvas)
dashboard.bind("<Visibility>", configure_dashboard_canvas)

frame = tkinter.Frame(dashboard_canvas, bg="white")

tkinter.Label(frame, text="SkillLink", font=("vintage", 36, "bold"), bg="white").grid(row=0, column=0, columnspan=2, pady=(30, 20))
button_font = ("Poppins", 16)
tkinter.Button(frame, text="Login", command=open_login, font=button_font, bg="black", fg="white", bd=0, padx=30, pady=15).grid(row=1, column=0, padx=20, pady=10, sticky="ew")
tkinter.Button(frame, text="Register", command=open_register, font=button_font, bg="white", fg="black", bd=3, highlightbackground="black", highlightthickness=1, padx=30, pady=15).grid(row=1, column=1, padx=20, pady=10, sticky="ew")

frame.grid_columnconfigure(0, weight=1)
frame.grid_columnconfigure(1, weight=1)
frame.grid_rowconfigure(0, weight=1)
frame.grid_rowconfigure(1, weight=1)

dashboard.mainloop()