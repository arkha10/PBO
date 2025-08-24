from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
import os
from PIL import Image, ImageTk

script_dir = os.path.dirname(os.path.abspath(__file__))
filepath = os.path.join(script_dir, 'data.xlsx')
BACKGROUND_IMAGE_PATH = os.path.join(script_dir, 'download1.png')

global_background_image_mp = None
global_background_photo_mp = None
current_window_width_mp = 0
current_window_height_mp = 0

def load_and_resize_background_mp(window, canvas):
    global global_background_image_mp, global_background_photo_mp, current_window_width_mp, current_window_height_mp

    canvas_width = canvas.winfo_width()
    canvas_height = canvas.winfo_height()

    if canvas_width == 0 or canvas_height == 0:
        window.after(100, lambda: load_and_resize_background_mp(window, canvas))
        return

    if global_background_image_mp is None:
        try:
            global_background_image_mp = Image.open(BACKGROUND_IMAGE_PATH)
        except FileNotFoundError:
            messagebox.showerror("Error", f"Background image not found: {BACKGROUND_IMAGE_PATH}")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load background image: {e}")
            return

    if canvas_width != current_window_width_mp or canvas_height != current_window_height_mp:
        resized_image = global_background_image_mp.resize((canvas_width, canvas_height), Image.LANCZOS)
        global_background_photo_mp = ImageTk.PhotoImage(resized_image)
        current_window_width_mp = canvas_width
        current_window_height_mp = canvas_height

    canvas.delete("background")
    canvas.create_image(0, 0, image=global_background_photo_mp, anchor="nw", tags="background")
    canvas.lower("background")


# Model
class UserDataModel:
    def __init__(self, filepath):
        self.filepath = filepath
        self._load_header()

    def _load_header(self):
        try:
            wb = load_workbook(self.filepath)
            sheet = wb.active
            self.header = [cell.value for cell in sheet[1] if cell.value is not None]
        except FileNotFoundError:
            self.header = []
        except Exception as e:
            raise e

    def get_all_users(self):
        try:
            wb = load_workbook(self.filepath)
            sheet = wb.active
            return list(sheet.iter_rows(min_row=2, values_only=True))
        except FileNotFoundError:
            return []
        except Exception as e:
            raise e

    def get_user_by_gmail(self, gmail): # Changed to get user by gmail
        users = self.get_all_users()
        gmail_col_idx = self.header.index("Gmail") if "Gmail" in self.header else -1
        if gmail_col_idx == -1:
            raise ValueError("Kolom 'Gmail' tidak ditemukan di header.")

        for i, user in enumerate(users):
            if user[gmail_col_idx].strip().lower() == gmail.strip().lower():
                return list(user), i + 2
        return None, None

    def update_user(self, original_gmail, updated_data): # Changed to update user by gmail
        try:
            wb = load_workbook(self.filepath)
            sheet = wb.active
            user_row_num = None
            
            gmail_col_idx = self.header.index("Gmail") if "Gmail" in self.header else -1
            if gmail_col_idx == -1:
                raise ValueError("Kolom 'Gmail' tidak ditemukan di header.")

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[gmail_col_idx].strip().lower() == original_gmail.strip().lower():
                    user_row_num = i
                    break

            if user_row_num:
                current_header_length = len(self.header)
                if len(updated_data) < current_header_length:
                    updated_data.extend([""] * (current_header_length - len(updated_data)))
                elif len(updated_data) > current_header_length:
                    updated_data = updated_data[:current_header_length]

                for col_num, value in enumerate(updated_data, start=1):
                    sheet.cell(row=user_row_num, column=col_num).value = value
                wb.save(self.filepath)
                return True
            return False
        except FileNotFoundError:
            raise FileNotFoundError("Error: data.xlsx not found!")
        except Exception as e:
            raise e

# View and Controller
class MainPage(Frame):
    def __init__(self, master, filepath, current_user=None): 
        super().__init__(master)
        self.master = master
        self.master.title("Main Page")
        self.master.state('zoomed')
        self.master.resizable(True, True)
        self.master.config(bg="#0077B5")

        self.current_user_gmail = current_user 
        self.filepath = filepath
        self.model = UserDataModel(self.filepath)

        self.background_canvas = Canvas(self.master, highlightthickness=0)
        self.background_canvas.pack(fill=BOTH, expand=True)

        self.main_wrapper_frame = Frame(self.background_canvas, bg="white", bd=2, relief=RIDGE)
        self.main_wrapper_frame_window = self.background_canvas.create_window((0, 0), window=self.main_wrapper_frame, anchor="nw")

        self.scrollable_content_container = Frame(self.main_wrapper_frame, bg="white")
        self.scrollable_content_container.pack(fill=BOTH, expand=1, pady=(10, 0), padx=10)

        self.content_canvas = Canvas(self.scrollable_content_container, bg="white", highlightthickness=0)
        self.content_canvas.pack(side=LEFT, fill=BOTH, expand=1)

        self.scrollbar = Scrollbar(self.scrollable_content_container, orient=VERTICAL, command=self.content_canvas.yview)
        self.scrollbar.pack(side=RIGHT, fill=Y)

        self.content_canvas.configure(yscrollcommand=self.scrollbar.set)

        self.content_frame = Frame(self.content_canvas, bg="white")
        self.content_window = self.content_canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        self.button_frame_inside = Frame(self.main_wrapper_frame, bg="white")
        self.button_frame_inside.pack(side=BOTTOM, fill=X, pady=10, padx=10)

        self.search_frame = Frame(self.main_wrapper_frame, bg="white", padx=10, pady=10)
        self.search_frame.pack(side=TOP, fill=X)

        Label(self.search_frame, text="Search:", font=("Poppins", 12), bg="white").pack(side=LEFT, padx=(0, 5))
        self.search_entry = Entry(self.search_frame, font=("Poppins", 12), width=30)
        self.search_entry.pack(side=LEFT, expand=True, fill=X, padx=(0, 10))
        self.search_button = Button(self.search_frame, text="Search", command=self.search_users,
                                     font=("Poppins", 12), bg="#00B0FF", fg="white", padx=10, pady=5)
        self.search_button.pack(side=LEFT, padx=(0, 5))
        self.show_all_button = Button(self.search_frame, text="Show All", command=self.load_and_display_data,
                                     font=("Poppins", 12), bg="#4CAF50", fg="white", padx=10, pady=5)
        self.show_all_button.pack(side=LEFT)


        def on_master_configure(event=None):
            load_and_resize_background_mp(self.master, self.background_canvas)

            master_width = self.master.winfo_width()
            master_height = self.master.winfo_height()
            
            wrapper_padding_x = master_width * 0.05
            wrapper_padding_y = master_height * 0.05
            
            wrapper_x = wrapper_padding_x
            wrapper_y = wrapper_padding_y
            wrapper_width = master_width - (2 * wrapper_padding_x)
            wrapper_height = master_height - (2 * wrapper_padding_y)

            self.background_canvas.coords(self.main_wrapper_frame_window, wrapper_x, wrapper_y)
            self.background_canvas.itemconfig(self.main_wrapper_frame_window, width=wrapper_width, height=wrapper_height)

            self.content_canvas.itemconfig(self.content_window, width=self.content_canvas.winfo_width())
            self.content_canvas.configure(scrollregion=self.content_canvas.bbox("all"))

        self.master.bind("<Configure>", on_master_configure)
        self.master.bind("<Visibility>", on_master_configure)
        self.content_frame.bind("<Configure>", lambda event: self.content_canvas.config(scrollregion=self.content_canvas.bbox("all")))

        self.edit_profile_button = Button(self.button_frame_inside, text="Edit Profile", command=self.edit_profile,
                                             font=("Poppins", 14), bg="#66a6ff", fg="white", padx=20, pady=10)
        self.edit_profile_button.pack(side=LEFT, padx=20, pady=10)

        self.logout_button = Button(self.button_frame_inside, text="Logout", command=self.logout,
                                         font=("Poppins", 14), bg="red", fg="white", padx=20, pady=10)
        self.logout_button.pack(side=RIGHT, padx=20, pady=10)

        self.load_and_display_data()

    def on_canvas_configure(self, event):
        canvas_width = event.width
        self.content_canvas.itemconfig(self.content_window, width=canvas_width)
        self.content_canvas.configure(scrollregion=self.content_canvas.bbox("all"))


    def load_and_display_data(self):
        try:
            for widget in self.content_frame.winfo_children():
                widget.destroy()

            users_data = self.model.get_all_users()
            header = self.model.header

            if not users_data:
                Label(self.content_frame, text="No user data found.",
                      font=("Poppins", 16), bg="white", fg="gray").pack(pady=20)
                return

            self.display_users(users_data, header)
            self.content_frame.update_idletasks()
            self.content_canvas.config(scrollregion=self.content_canvas.bbox("all"))
        except FileNotFoundError:
            Label(self.content_frame, text="Error: data.xlsx not found!",
                  font=("Poppins", 16), bg="white", fg="red").pack(pady=20)
        except Exception as e:
            Label(self.content_frame, text=f"An error occurred: {e}",
                  font=("Poppins", 16), bg="white", fg="red").pack(pady=20)

    def search_users(self):
        query = self.search_entry.get().strip().lower()
        if not query:
            messagebox.showinfo("Search", "Please enter a word.")
            self.load_and_display_data()
            return
        try:
            all_users = self.model.get_all_users()
            header = self.model.header
            filtered_users = []
            for user in all_users:
                found = False
                for i, value in enumerate(user):
                    if header[i] == "Password":
                        continue
                    if header[i] == "Project":
                        if value:
                            projects = str(value).split('\n')
                            if any(query in p.strip().lower() for p in projects if p.strip()):
                                found = True
                                break
                        continue

                    if str(value).strip().lower().find(query) != -1:
                        found = True
                        break
                if found:
                    filtered_users.append(user)

            for widget in self.content_frame.winfo_children():
                widget.destroy()

            if not filtered_users:
                Label(self.content_frame, text=f"No users found matching '{query}'.",
                      font=("Poppins", 16), bg="white", fg="gray").pack(pady=20)
            else:
                self.display_users(filtered_users, header)

            self.content_frame.update_idletasks()
            self.content_canvas.config(scrollregion=self.content_canvas.bbox("all"))

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during search: {e}")

    def display_users(self, users_data, header):
        gmail_index = header.index("Gmail") if "Gmail" in header else -1 

        for user in users_data:
            user_frame = Frame(self.content_frame, bg="#f0f0f0", bd=2, relief=SOLID, padx=15, pady=10)
            user_frame.pack(pady=10, padx=20, fill="x")

            password_index = header.index("Password") if "Password" in header else -1
            project_index = header.index("Project") if "Project" in header else -1
            

            display_header = []
            display_user_data = []

            for j, h in enumerate(header):
                if h == "Password":
                    continue
                display_header.append(h)
                display_user_data.append(user[j])

            # Check if this is the current user's profile based on Gmail
            if gmail_index != -1 and user[gmail_index].strip().lower() == self.current_user_gmail.strip().lower():
                user_frame.config(bg="#e0f7fa", highlightbackground="#26c6da", highlightthickness=2)
                Label(user_frame, text="Your Profile", font=("Poppins", 18, "bold"),
                      bg="#e0f7fa", fg="#0097a7").pack(anchor="w", pady=(0, 10))

            for j, value in enumerate(display_user_data):
                field = display_header[j]
                if field == "Project":
                    Label(user_frame, text=f"{field}:", font=("Poppins", 12, "bold"),
                          bg=user_frame["bg"], anchor="w").pack(fill="x", pady=(5, 0))
                    projects = str(value).split('\n') if value else []
                    if projects:
                        for project in projects:
                            if project.strip():
                                Label(user_frame, text=f"  - {project.strip()}", font=("Poppins", 12),
                                      bg=user_frame["bg"], anchor="w", justify=LEFT).pack(fill="x", padx=10)
                    else:
                        Label(user_frame, text="  (No projects yet)", font=("Poppins", 12, "italic"),
                              bg=user_frame["bg"], anchor="w", justify=LEFT).pack(fill="x", padx=10)
                else:
                    label_text = f"{field}: {value}"
                    Label(user_frame, text=label_text, font=("Poppins", 12),
                          bg=user_frame["bg"], anchor="w").pack(fill="x", pady=2)

        self.content_frame.update_idletasks()
        self.content_canvas.configure(scrollregion=self.content_canvas.bbox("all"))


    def edit_profile(self):
        if not self.current_user_gmail:
            messagebox.showerror("Error", "Tidak dapat mengedit profil. Silakan login kembali.")
            return

        user_data, _ = self.model.get_user_by_gmail(self.current_user_gmail) # Get user data by gmail
        header = self.model.header

        if not user_data:
            messagebox.showerror("Error", "Data pengguna tidak ditemukan")
            return

        edit_window = Toplevel(self.master)
        edit_window.title("Edit Profile")
        edit_window.state('zoomed')
        edit_window.resizable(True, True)
        edit_window.config(bg="#0077B5")

        edit_bg_canvas = Canvas(edit_window, highlightthickness=0)
        edit_bg_canvas.pack(fill=BOTH, expand=True)

        edit_form_wrapper_frame = Frame(edit_bg_canvas, bg="white", bd=2, relief=RIDGE)
        edit_form_wrapper_window = edit_bg_canvas.create_window((0, 0), window=edit_form_wrapper_frame, anchor="nw")

        def configure_edit_canvas(event=None):
            load_and_resize_background_mp(edit_window, edit_bg_canvas)
            
            canvas_width = edit_bg_canvas.winfo_width()
            canvas_height = edit_bg_canvas.winfo_height()
            
            edit_form_wrapper_frame.update_idletasks()
            form_frame_width = edit_form_wrapper_frame.winfo_width()
            form_frame_height = edit_form_wrapper_frame.winfo_height()
            
            x_pos = (canvas_width - form_frame_width) / 2
            y_pos = (canvas_height - form_frame_height) / 2
            
            if form_frame_width > 0 and form_frame_height > 0:
                edit_bg_canvas.coords(edit_form_wrapper_window, x_pos, y_pos)
                edit_bg_canvas.itemconfig(edit_form_wrapper_window, anchor="nw")
            else:
                edit_bg_canvas.coords(edit_form_wrapper_window, canvas_width/2, canvas_height/2)
                edit_bg_canvas.itemconfig(edit_form_wrapper_window, anchor="center")

        edit_bg_canvas.bind("<Configure>", configure_edit_canvas)
        edit_window.bind("<Visibility>", configure_edit_canvas)
        edit_form_wrapper_frame.bind("<Configure>", configure_edit_canvas)

        Label(edit_form_wrapper_frame, text="Edit Profile", font=("Poppins", 24, "bold"), bg="white").pack(pady=20)

        entries = {}
        text_widgets = {}

        original_gmail_from_data = user_data[header.index("Gmail")].strip().lower() if "Gmail" in header else ""


        for i, field in enumerate(header):
            # We will handle 'Password' as a special case for its Entry widget, but still include it in the iteration
            # for data retrieval and potential update.
            Label(edit_form_wrapper_frame, text=field + ":", bg="white", font=("Poppins", 12)).pack(pady=(10, 0))
            if field == "Project":
                project_text = Text(edit_form_wrapper_frame, height=5, width=40, font=("Poppins", 12))
                project_text.insert("1.0", str(user_data[i]) if user_data[i] is not None else "")
                project_text.pack(padx=20, pady=5, fill=X)
                text_widgets[field] = project_text
            else:
                entry_var = StringVar(value=user_data[i] if user_data[i] is not None else "")
                ent = Entry(edit_form_wrapper_frame, textvariable=entry_var, show="*" if field == "Password" else None, font=("Poppins", 12))
                ent.pack(padx=20, pady=5, fill=X)
                entries[field] = entry_var

        def save_changes():
            updated_values = []
            new_gmail_for_check = ""

            mandatory_fields = ["Nama", "Gmail", "Pendidikan Terakhir", "Nomor Telepon", "Password", "Bidang yang dikuasai"]

            for field in header:
                if field == "Project":
                    project_input = text_widgets[field].get("1.0", END).strip()
                    projects_list = [p.strip() for p in project_input.split('\n') if p.strip()]
                    updated_values.append('\n'.join(projects_list))
                else:
                    val = entries[field].get().strip()
                    if field in mandatory_fields and val == "":
                        messagebox.showerror("Error", f"Kolom '{field}' harus diisi!")
                        return
                    updated_values.append(val)
                    if field == "Gmail":
                        new_gmail_for_check = val.lower()

            if new_gmail_for_check != original_gmail_from_data:
                wb = load_workbook(self.filepath)
                sheet = wb.active
                
                gmail_col_idx = header.index("Gmail")
                
                # Exclude the current user's old gmail from the existing check
                existing_gmails = [
                    str(row[gmail_col_idx]).strip().lower() 
                    for row in sheet.iter_rows(min_row=2, values_only=True)
                    if str(row[gmail_col_idx]).strip().lower() != original_gmail_from_data
                ]

                if new_gmail_for_check in existing_gmails:
                    messagebox.showerror("Error", "Gmail ini sudah terdaftar oleh pengguna lain! Gunakan gmail yang berbeda.")
                    return
            
            try:
                # Update using the original Gmail to find the user
                if self.model.update_user(original_gmail_from_data, updated_values):
                    # If Gmail was changed, update current_user_gmail
                    if new_gmail_for_check != original_gmail_from_data:
                        self.current_user_gmail = new_gmail_for_check 
                    messagebox.showinfo("Sukses", "Profil berhasil diperbarui!")
                    edit_window.destroy()
                    self.load_and_display_data() # Reload data to reflect changes
                else:
                    messagebox.showerror("Error", "Gagal memperbarui profil.")
            except FileNotFoundError as e:
                messagebox.showerror("Error", str(e))
            except Exception as e:
                messagebox.showerror("Error", f"Terjadi kesalahan saat menyimpan: {e}")

        Button(edit_form_wrapper_frame, text="Simpan", command=save_changes,
               font=("Poppins", 14), bg="#66a6ff", fg="white",
               bd=0, padx=20, pady=10).pack(pady=15)
        Button(edit_form_wrapper_frame, text="Batal", command=edit_window.destroy,
               font=("Poppins", 14), bg="gray", fg="white",
               bd=0, padx=20, pady=10).pack()

    def logout(self):
        import sys
        import subprocess
        import os

        login_script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'login.py')

        self.master.destroy()
        subprocess.Popen([sys.executable, login_script_path])

if __name__ == '__main__':
    root = Tk()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(script_dir, 'data.xlsx') 
    main_page = MainPage(root, filepath, current_user="test@example.com") 
    root.mainloop()